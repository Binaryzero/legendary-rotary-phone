import logging
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional
import re

import requests
from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s: %(message)s")

MITRE_BASE = "https://raw.githubusercontent.com/CVEProject/cvelistV5/main/cves"

# Precompiled regex for validating CVE identifiers
cve_pattern = re.compile(r"^CVE-(?P<year>\d{4})-(?P<number>\d{4,})$")


@dataclass
class CveMetadata:
    """Container for parsed CVE information."""

    description: str = ""
    cvss: str = ""
    vector: str = ""
    cwe: str = ""
    exploit: str = ""
    exploit_refs: str = ""
    fix_version: str = ""
    mitigations: str = ""
    affected: str = ""
    references: str = ""

def fetch_cve(cve_id: str, session: Optional[requests.Session] = None) -> Optional[dict]:
    """Retrieve a CVE JSON record from the MITRE repository.

    Parameters
    ----------
    cve_id:
        The CVE identifier, e.g. ``"CVE-2023-1234"``.
    session:
        Optional :class:`requests.Session` to reuse for HTTP calls.

    Returns
    -------
    dict | None
        Parsed JSON content if found, otherwise ``None`` when the request fails.
    """

    match = cve_pattern.fullmatch(cve_id)
    if not match:
        logging.warning("Invalid CVE ID format: %s", cve_id)
        return None

    year = match.group("year")
    number_str = match.group("number")
    bucket = int(number_str) // 1000

    url = f"{MITRE_BASE}/{year}/{bucket}xxx/{cve_id}.json"
    sess = session or requests.Session()
    try:
        response = sess.get(url, timeout=10)
        response.raise_for_status()
    except requests.RequestException as exc:
        logging.warning("Failed to fetch %s: %s", cve_id, exc)
        return None
    finally:
        if session is None:
            sess.close()
    return response.json()

def parse_cve(cve_json: dict) -> CveMetadata:
    """Extract relevant metadata from a CVE JSON document.

    Args:
        cve_json (dict): JSON structure as returned by :func:`fetch_cve`.

    Returns
    -------
    CveMetadata
        Structured metadata extracted from the CVE document.
    """
    cna = cve_json.get("containers", {}).get("cna", {})
    descriptions_list = cna.get("descriptions", [])
    desc = descriptions_list[0].get("value", "") if descriptions_list else ""

    metrics_list = cna.get("metrics", [])
    cvss_data = metrics_list[0].get("cvssV3_1", {}) if metrics_list else {}
    base_score_val = cvss_data.get("baseScore", "")
    cvss = str(base_score_val) if base_score_val != "" else ""
    vector = cvss_data.get("vectorString", "")
    cwe_items = []
    for pt in cna.get("problemTypes", []):
        for desc_entry in pt.get("descriptions", []):
            val = desc_entry.get("description") or desc_entry.get("value")
            if val:
                cwe_items.append(val)
    cwe = ", ".join(cwe_items)
    references = cna.get("references", [])
    exploits = [r["url"] for r in references if "exploit-db.com" in r["url"] or "github.com" in r["url"]]
    exploit_flag = "Yes" if exploits else "No"
    fix_version = ""
    mitigations = []
    affected = []
    for r in references:
        url = r.get("url", "")
        if "advisories" in url or "bulletin" in url:
            mitigations.append(url)
        if "upgrade" in r.get("tags", []):
            fix_version = r.get("url", "")
    for a in cna.get("affected", []):
        vendor = a.get("vendor", "")
        product = a.get("product", "")
        versions = ", ".join(v.get("version", "") for v in a.get("versions", []))
        item = " ".join(i for i in [vendor, product, versions] if i)
        if item:
            affected.append(item)
    return CveMetadata(
        description=desc,
        cvss=cvss,
        vector=vector,
        cwe=cwe,
        exploit=exploit_flag,
        exploit_refs=", ".join(exploits),
        fix_version=fix_version,
        mitigations=", ".join(mitigations),
        affected="; ".join(affected),
        references=", ".join(r.get("url", "") for r in references),
    )


def create_report(cve_id: str, meta: CveMetadata, template_path: Path, output_dir: Path) -> None:
    """Generate a Word report based on the template.

    Parameters
    ----------
    cve_id: str
        Identifier for the CVE.
    meta:
        Parsed metadata dataclass returned by :func:`parse_cve`.
    template_path:
        Path to the Word template file.
    output_dir:
        Directory where the generated report will be stored.
    """
    if not template_path.exists():
        logging.error("%s not found", template_path)
        return
    doc = Document(template_path)

    for p in doc.paragraphs:
        txt = p.text.strip()
        if txt.startswith("CVE ID:"):
            p.text = f"CVE ID: {cve_id}"
        elif txt == "Brief summary of the vulnerability.":
            p.text = meta.description
        elif txt == "List of affected products or environments.":
            p.text = meta.affected
        elif txt.startswith("Describe exploit vectors"):
            exp = (
                f"CVSS: {meta.cvss} ({meta.vector})\n"
                f"CWE: {meta.cwe}\n"
                f"Exploit Available: {meta.exploit} {meta.exploit_refs}"
            )
            p.text = exp
        elif txt.startswith("Document any known fixes"):
            mitigations = []
            if meta.fix_version:
                mitigations.append(f"Fix: {meta.fix_version}")
            if meta.mitigations:
                mitigations.append(f"Mitigations: {meta.mitigations}")
            p.text = "\n".join(mitigations)
        elif txt.startswith("List external references"):
            p.text = meta.references

    output_dir.mkdir(exist_ok=True)
    doc.save(output_dir / f"{cve_id}.docx")

def main(
    input_file: Path = Path("cves.txt"),
    excel_out: Path = Path("CVE_Results.xlsx"),
    template_path: Path = Path("CVE_Report_Template.docx"),
    report_dir: Path = Path("reports"),
    write_reports: bool = True,
) -> None:
    """Read CVE IDs, fetch their metadata and save results to Excel."""
    if not input_file.exists():
        logging.error("Missing %s input file.", input_file)
        return
    cve_ids = [line.strip() for line in input_file.read_text().splitlines() if line.strip()]
    wb = Workbook()
    ws = wb.active
    ws.title = f"Batch_{datetime.now().strftime('%Y%m%d_%H%M')}"
    ws.append([
        "CVE ID",
        "Description",
        "CVSS",
        "Vector",
        "CWE",
        "Exploit",
        "ExploitRefs",
        "FixVersion",
        "Mitigations",
        "Affected",
        "References",
    ])
    ws.freeze_panes = "A2"
    header = ws[1]
    for cell in header:
        cell.font = Font(bold=True)

    with requests.Session() as session:
        for cve_id in cve_ids:
            logging.info("Processing %s", cve_id)
            data = fetch_cve(cve_id, session)
            if not data:
                continue
            parsed = parse_cve(data)
            ws.append([
                cve_id,
                parsed.description,
                parsed.cvss,
                parsed.vector,
                parsed.cwe,
                parsed.exploit,
                parsed.exploit_refs,
                parsed.fix_version,
                parsed.mitigations,
                parsed.affected,
                parsed.references,
            ])
            if write_reports:
                create_report(cve_id, parsed, template_path, report_dir)

    for idx, column in enumerate(ws.iter_cols(max_col=ws.max_column, max_row=ws.max_row), start=1):
        length = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[get_column_letter(idx)].width = max(length + 2, 10)

    wb.save(excel_out)
    logging.info("%s created.", excel_out)

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Fetch CVE metadata from MITRE")
    parser.add_argument("input", nargs="?", default="cves.txt", help="Path to file containing CVE IDs")
    parser.add_argument("--excel", default="CVE_Results.xlsx", help="Output Excel file")
    parser.add_argument("--template", default="CVE_Report_Template.docx", help="Word template path")
    parser.add_argument("--reports", default="reports", help="Directory for Word reports")
    parser.add_argument("--skip-docs", action="store_true", help="Skip Word report generation")
    args = parser.parse_args()

    main(
        Path(args.input),
        Path(args.excel),
        Path(args.template),
        Path(args.reports),
        write_reports=not args.skip_docs,
    )
